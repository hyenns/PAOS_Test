from __future__ import annotations

import base64
import io
import json
import os
import re
from pathlib import Path

import pandas as pd
from flask import Flask, Response, jsonify, request, send_from_directory, stream_with_context

from backend.iros_crawler import IROS_RESULT_COLUMNS, run_iros_crawler_events
from backend.saramin_crawler import SARAMIN_RESULT_COLUMNS, get_saramin_result_columns, run_saramin_crawler_events

BASE_DIR = Path(__file__).resolve().parent

app = Flask(__name__)
# JSON 응답의 키 순서가 자동 정렬되면 화면 미리보기 열 순서가 엑셀과 달라질 수 있어 비활성화합니다.
app.json.sort_keys = False
app.config["MAX_CONTENT_LENGTH"] = 30 * 1024 * 1024  # 30MB


def excel_column_letter(index: int) -> str:
    index += 1
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def read_search_values(
    file_bytes: bytes,
    sheet_name: str,
    header_mode: str,
    column_index: int,
    column_letter: str = "",
) -> list[str]:
    """선택한 엑셀 열의 검색값을 읽습니다.

    프론트에서 보이는 열 문자와 실제 엑셀 열이 어긋나지 않도록,
    가능하면 column_letter(예: AA)를 기준으로 openpyxl에서 직접 읽습니다.
    엑셀 맨 앞에 빈 열이 있어 SheetJS 미리보기와 pandas 열 인덱스가 다르게 잡히는 경우를 방지합니다.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        if column_letter:
            column_number = column_index_from_string(str(column_letter).strip().upper())
        else:
            column_number = column_index + 1

        if column_number < 1 or column_number > max(ws.max_column or 1, column_number):
            raise ValueError("선택한 열 번호가 엑셀 범위를 벗어났습니다.")

        start_row = 2 if header_mode == "header" else 1
        values: list[str] = []

        for row_number in range(start_row, (ws.max_row or 0) + 1):
            # 완전히 빈 행은 건너뜁니다.
            row_values = [
                ws.cell(row=row_number, column=col).value
                for col in range(1, (ws.max_column or column_number) + 1)
            ]
            if all(value is None or str(value).strip() == "" for value in row_values):
                continue

            value = ws.cell(row=row_number, column=column_number).value
            if value is None:
                continue

            cleaned = str(value).strip()
            if not cleaned or cleaned.lower() == "nan":
                continue
            values.append(cleaned)

        return values
    finally:
        wb.close()


def apply_result_excel_style(worksheet) -> None:
    """크롤링 결과 엑셀 파일의 가독성을 높이는 공통 서식입니다."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    max_row = worksheet.max_row or 1
    max_column = worksheet.max_column or 1

    # 첫 행 고정 + 필터
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    # 요청 반영: 헤더는 초록색이 아닌 회색으로 표시
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(color="000000", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )
    body_alignment = Alignment(vertical="center", wrap_text=False)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.row_dimensions[1].height = 22

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.alignment = body_alignment
            cell.border = thin_border

    # 내용 길이에 맞춰 열 너비 조정
    for column_index in range(1, max_column + 1):
        column_letter = get_column_letter(column_index)
        header_value = worksheet.cell(row=1, column=column_index).value
        header_text = str(header_value or "")
        max_length = len(header_text)

        for cell in worksheet[column_letter]:
            value = cell.value
            if value is None:
                continue
            value_text = str(value)
            # 빈 문자열인 셀이 있어도 오류가 나지 않도록 안전하게 계산합니다.
            parts = value_text.splitlines() or [value_text]
            max_length = max(max_length, max(len(part) for part in parts))

        width = max_length + 3
        if any(keyword in header_text for keyword in ["주소", "소재지", "사업내용"]):
            width = min(max(width, 18), 55)
        elif "URL" in header_text or "홈페이지" in header_text:
            width = min(max(width, 18), 50)
        elif any(keyword in header_text for keyword in ["법인등록번호", "사업자등록번호"]):
            width = min(max(width, 18), 24)
        else:
            width = min(max(width, 10), 28)

        worksheet.column_dimensions[column_letter].width = width


# 회사 내부 표준 주소 표기: 도/광역시 전체 명칭을 약어로 바꾸되, 뒤 주소는 그대로 유지합니다.
PROVINCE_REPLACEMENTS = [
    ("서울특별시", "서울"),
    ("부산광역시", "부산"),
    ("대구광역시", "대구"),
    ("인천광역시", "인천"),
    ("광주광역시", "광주"),
    ("대전광역시", "대전"),
    ("울산광역시", "울산"),
    ("세종특별자치시", "세종"),
    ("경기도", "경기"),
    ("강원특별자치도", "강원"),
    ("강원도", "강원"),
    ("충청북도", "충북"),
    ("충청남도", "충남"),
    ("전북특별자치도", "전북"),
    ("전라북도", "전북"),
    ("전라남도", "전남"),
    ("경상북도", "경북"),
    ("경상남도", "경남"),
    ("제주특별자치도", "제주"),
    ("제주도", "제주"),
]

MONEY_COLUMNS = ["매출액", "영업이익", "당기순이익", "자본금", "평균연봉"]


def clean_text_value(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def standardize_province_in_address(value: object) -> str:
    """시도명만 회사 표준 약어로 바꾸고, 시군구 이하 주소는 그대로 둡니다."""
    text = clean_text_value(value)
    if not text:
        return ""

    for full_name, short_name in PROVINCE_REPLACEMENTS:
        match = re.match(rf"^\s*{re.escape(full_name)}\s*", text)
        if match:
            rest = text[match.end():].strip()
            return f"{short_name} {rest}".strip()

    return text


def split_korean_english_name(value: object) -> tuple[str, str]:
    text = clean_text_value(value)
    if not text:
        return "", ""

    # 예: 큐픽스(Cupix, Inc.) -> 큐픽스 / Cupix, Inc.
    match = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", text)
    if match:
        korean_name = clean_text_value(match.group(1))
        english_name = clean_text_value(match.group(2))
        return korean_name, english_name

    return text, ""


def remove_company_prefix(value: object) -> str:
    text = clean_text_value(value)
    if not text:
        return ""
    text = text.replace("㈜", "")
    text = re.sub(r"^\s*(?:\(주\)|（주）|주식회사)\s*", "", text)
    text = re.sub(r"\s*(?:\(주\)|（주）|주식회사)\s*$", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_establish_date(value: object) -> str:
    text = clean_text_value(value)
    if not text:
        return ""

    match = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", text)
    if match:
        year, month, day = map(int, match.groups())
        return f"{year:04d}-{month:02d}-{day:02d}"

    match = re.search(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", text)
    if match:
        year, month, day = map(int, match.groups())
        return f"{year:04d}-{month:02d}-{day:02d}"

    return text


def clean_employee_count(value: object) -> str:
    text = clean_text_value(value)
    if not text:
        return ""
    number_text = re.sub(r"[^0-9]", "", text)
    if not number_text:
        return ""
    return f"{int(number_text):,}"


def clean_homepage_url(value: object) -> str:
    text = clean_text_value(value)
    if not text:
        return ""
    text = re.sub(r"^https?://", "", text, flags=re.I)
    return text.rstrip("/").strip()


def format_million_integer_without_rounding(value) -> str:
    try:
        from decimal import Decimal, ROUND_DOWN
        decimal_value = value if isinstance(value, Decimal) else Decimal(str(value))
    except Exception:
        return ""

    # 백만원 미만 소수점은 반올림하지 않고 버립니다. 예: 17,383.56 -> 17,383
    integer_value = decimal_value.quantize(Decimal("1"), rounding=ROUND_DOWN)
    return f"{int(integer_value):,}"


def parse_korean_money_to_million(value: object) -> str:
    """사람인 금액 텍스트를 백만원 단위 정수 문자열로 변환합니다. 소수점은 반올림 없이 버립니다."""
    from decimal import Decimal

    text = clean_text_value(value)
    if not text or text in {"-", "--", "없음"}:
        return ""

    # 전년대비/순위 등 부가 문구가 섞여 있어도 단위 숫자만 계산합니다.
    total = Decimal("0")
    found_unit = False

    unit_multipliers = [
        ("조", Decimal("1000000")),
        ("억", Decimal("100")),
        ("만원", Decimal("0.01")),
    ]

    for unit, multiplier in unit_multipliers:
        match = re.search(rf"([-+]?\d[\d,]*(?:\.\d+)?)\s*{unit}", text)
        if match:
            number = Decimal(match.group(1).replace(",", ""))
            total += number * multiplier
            found_unit = True

    if not found_unit:
        match = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?", text)
        if not match:
            return ""
        # 이미 숫자만 있는 값은 백만원 단위 숫자로 보고 쉼표만 정리합니다.
        total = Decimal(match.group(0).replace(",", ""))

    return format_million_integer_without_rounding(total)


def build_iros_clean_rows(
    rows: list[dict],
    options: dict[str, bool],
    columns: list[str] | None = None,
) -> tuple[list[dict], list[str]]:
    """등기소 결과를 회사 형식으로 정제합니다.

    상호검색은 기존 고정 열을 사용하고, 등록번호검색은 인터넷등기소 결과표의
    실제 열 이름을 그대로 가져오므로 columns를 기준으로 동적으로 정제합니다.
    """
    split_name = options.get("split_name", False)
    remove_reg_hyphen = options.get("remove_reg_hyphen", False)
    standardize_address = options.get("standardize_address", False)
    source_columns = list(columns or IROS_RESULT_COLUMNS)

    name_columns = {"상호(명칭)", "상호", "법인명", "상호명"}
    reg_columns = {"법인등록번호", "등록번호", "법인번호"}
    address_columns = {"본점소재지", "본점", "본점주소", "소재지"}

    clean_columns: list[str] = []
    for column in source_columns:
        if column in name_columns and split_name:
            clean_columns.extend(["상호_국문", "상호_영문"])
        else:
            clean_columns.append(column)

    # 중복된 정제 열 이름이 생기는 경우 한 번만 유지
    clean_columns = list(dict.fromkeys(clean_columns))

    clean_rows: list[dict] = []
    for row in rows:
        clean_row: dict = {}
        for column in source_columns:
            value = row.get(column, "")

            if column in name_columns and split_name:
                korean_name, english_name = split_korean_english_name(value)
                clean_row["상호_국문"] = korean_name
                clean_row["상호_영문"] = english_name
                continue

            if column in reg_columns and remove_reg_hyphen:
                value = clean_text_value(value).replace("-", "")
            elif column in address_columns and standardize_address:
                value = standardize_province_in_address(value)

            clean_row[column] = value
        clean_rows.append(clean_row)

    return clean_rows, clean_columns


def build_saramin_clean_rows(rows: list[dict], columns: list[str], options: dict[str, bool]) -> tuple[list[dict], list[str]]:
    clean_company = options.get("company_name", False)
    clean_date = options.get("establish_date", False)
    clean_employee = options.get("employee_count", False)
    clean_homepage = options.get("homepage", False)
    clean_money = options.get("money_to_million", False)

    clean_rows: list[dict] = []
    for row in rows:
        clean_row: dict = {}
        for column in columns:
            value = row.get(column, "")

            if column == "회사명" and clean_company:
                value = remove_company_prefix(value)
            elif column == "설립일" and clean_date:
                value = normalize_establish_date(value)
            elif column == "사원수" and clean_employee:
                value = clean_employee_count(value)
            elif column == "홈페이지" and clean_homepage:
                value = clean_homepage_url(value)
            elif clean_money and any(keyword in column for keyword in MONEY_COLUMNS):
                value = parse_korean_money_to_million(value)

            clean_row[column] = value
        clean_rows.append(clean_row)

    return clean_rows, columns


def dataframe_to_excel_base64(
    rows: list[dict],
    columns: list[str] | None = None,
    *,
    clean_rows: list[dict] | None = None,
    clean_columns: list[str] | None = None,
) -> str:
    df = pd.DataFrame(rows, columns=columns) if columns else pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="결과")
        apply_result_excel_style(writer.sheets["결과"])

        if clean_rows is not None:
            clean_df = pd.DataFrame(clean_rows, columns=clean_columns) if clean_columns else pd.DataFrame(clean_rows)
            clean_df.to_excel(writer, index=False, sheet_name="정제결과")
            apply_result_excel_style(writer.sheets["정제결과"])

    output.seek(0)
    return base64.b64encode(output.read()).decode("utf-8")


@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/<path:path>")
def static_files(path: str):
    target = BASE_DIR / path
    if target.exists() and target.is_file():
        return send_from_directory(BASE_DIR, path)
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/api/health")
def health():
    return jsonify({"ok": True, "app": "WorkLab PAOS", "mode": "team-test"})


@app.route("/api/iros/run", methods=["POST"])
def api_iros_run():
    uploaded = request.files.get("file")
    if not uploaded:
        return jsonify({"ok": False, "message": "엑셀 파일이 업로드되지 않았습니다."}), 400

    sheet_name = request.form.get("sheet_name", "")
    header_mode = request.form.get("header_mode", "header")
    column_letter = request.form.get("column_letter", "").strip().upper()
    search_mode = request.form.get("search_mode", "company").strip().lower()
    if search_mode not in {"company", "registration"}:
        return jsonify({"ok": False, "message": "검색 방식이 올바르지 않습니다."}), 400

    try:
        column_index = int(request.form.get("column_index", "0"))
    except ValueError:
        return jsonify({"ok": False, "message": "열 번호가 올바르지 않습니다."}), 400

    try:
        file_bytes = uploaded.read()
        search_values = read_search_values(file_bytes, sheet_name, header_mode, column_index, column_letter)
    except Exception as e:
        return jsonify({"ok": False, "message": f"엑셀 파일을 읽을 수 없습니다: {str(e)}"}), 400

    if not search_values:
        target_name = "법인등록번호" if search_mode == "registration" else "회사명"
        return jsonify({"ok": False, "message": f"선택한 열에 크롤링할 {target_name}이 없습니다."}), 400

    include_closed_records = request.form.get("include_closed_records", "false").lower() == "true"
    include_erased_names = request.form.get("include_erased_names", "false").lower() == "true"

    clean_iros_enabled = request.form.get("clean_iros_enabled", "false").lower() == "true"
    iros_clean_options = {
        "split_name": request.form.get("clean_iros_split_name", "false").lower() == "true",
        "remove_reg_hyphen": request.form.get("clean_iros_remove_reg_hyphen", "false").lower() == "true",
        "standardize_address": request.form.get("clean_iros_standardize_address", "false").lower() == "true",
    }

    # 기본값은 Chrome 창을 띄우지 않는 백그라운드 모드입니다.
    # 추후 문제가 생기면 프론트에서 headless=false를 넘기거나 환경변수/코드로 조정할 수 있습니다.
    headless = request.form.get("headless", "true").lower() != "false"

    def ndjson(event: dict) -> str:
        return json.dumps(event, ensure_ascii=False) + "\n"

    @stream_with_context
    def generate():
        for event in run_iros_crawler_events(
            search_values,
            search_mode=search_mode,
            include_closed_records=include_closed_records,
            include_erased_names=include_erased_names,
            headless=headless,
        ):
            if event.get("type") == "complete":
                results = event.get("results", [])
                clean_rows = None
                clean_columns = None
                clean_applied = any(iros_clean_options.values())
                clean_sheet_added = False
                crawler_columns = event.get("columns") or IROS_RESULT_COLUMNS
                output_rows = results
                output_columns = crawler_columns
                display_rows = results
                display_columns = crawler_columns

                if clean_applied:
                    clean_rows, clean_columns = build_iros_clean_rows(
                        results, iros_clean_options, crawler_columns
                    )
                    display_rows = clean_rows
                    display_columns = clean_columns
                    if clean_iros_enabled:
                        # 체크한 경우: 원본 결과 시트 + 정제결과 시트를 함께 저장합니다.
                        clean_sheet_added = True
                    else:
                        # 체크하지 않은 경우: 선택한 정제를 결과 시트에 바로 적용하여 한 시트로 저장합니다.
                        output_rows = clean_rows
                        output_columns = clean_columns
                        clean_rows = None
                        clean_columns = None

                excel_base64 = dataframe_to_excel_base64(
                    output_rows,
                    columns=output_columns,
                    clean_rows=clean_rows,
                    clean_columns=clean_columns,
                )

                status_columns = ["등기상태", "상호말소상태", "주말 여부"]
                status_has_value = any(
                    str(row.get(col, "")).strip()
                    for row in results
                    for col in status_columns
                )

                event.update({
                    "ok": True,
                    "search_mode": search_mode,
                    "total_input": len(search_values),
                    "total_result": len(results),
                    "status_has_value": status_has_value,
                    "columns": display_columns,
                    "results": display_rows,
                    "clean_applied": clean_applied,
                    "clean_sheet_added": clean_sheet_added,
                    "excel_base64": excel_base64,
                    "filename": (
                        "등기소_등록번호검색_결과.xlsx"
                        if search_mode == "registration"
                        else "등기소_상호검색_결과.xlsx"
                    ),
                })

            yield ndjson(event)

    return Response(
        generate(),
        mimetype="application/x-ndjson; charset=utf-8",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.route("/api/saramin/run", methods=["POST"])
def api_saramin_run():
    uploaded = request.files.get("file")
    if not uploaded:
        return jsonify({"ok": False, "message": "엑셀 파일이 업로드되지 않았습니다."}), 400

    sheet_name = request.form.get("sheet_name", "")
    header_mode = request.form.get("header_mode", "header")
    column_letter = request.form.get("column_letter", "").strip().upper()

    try:
        column_index = int(request.form.get("column_index", "0"))
    except ValueError:
        return jsonify({"ok": False, "message": "열 번호가 올바르지 않습니다."}), 400

    try:
        max_results_per_keyword = int(request.form.get("max_results_per_keyword", "5"))
        max_results_per_keyword = max(1, min(max_results_per_keyword, 5))
    except ValueError:
        max_results_per_keyword = 5

    try:
        file_bytes = uploaded.read()
        search_values = read_search_values(file_bytes, sheet_name, header_mode, column_index, column_letter)
    except Exception as e:
        return jsonify({"ok": False, "message": f"엑셀 파일을 읽을 수 없습니다: {str(e)}"}), 400

    if not search_values:
        return jsonify({"ok": False, "message": "선택한 열에 크롤링할 회사명이 없습니다."}), 400

    headless = request.form.get("headless", "true").lower() != "false"
    collect_finance = request.form.get("collect_finance", "false").lower() == "true"
    result_columns = get_saramin_result_columns(collect_finance)

    clean_saramin_enabled = request.form.get("clean_saramin_enabled", "false").lower() == "true"
    saramin_clean_options = {
        "company_name": request.form.get("clean_saramin_company_name", "false").lower() == "true",
        "establish_date": request.form.get("clean_saramin_establish_date", "false").lower() == "true",
        "employee_count": request.form.get("clean_saramin_employee_count", "false").lower() == "true",
        "homepage": request.form.get("clean_saramin_homepage", "false").lower() == "true",
        "money_to_million": request.form.get("clean_saramin_money_to_million", "false").lower() == "true",
    }

    def ndjson(event: dict) -> str:
        return json.dumps(event, ensure_ascii=False) + "\n"

    @stream_with_context
    def generate():
        for event in run_saramin_crawler_events(
            search_values,
            max_results_per_keyword=max_results_per_keyword,
            collect_finance=collect_finance,
            headless=headless,
        ):
            if event.get("type") == "complete":
                results = event.get("results", [])
                clean_rows = None
                clean_columns = None
                clean_applied = any(saramin_clean_options.values())
                clean_sheet_added = False
                output_rows = results
                output_columns = result_columns
                display_rows = results
                display_columns = result_columns

                if clean_applied:
                    clean_rows, clean_columns = build_saramin_clean_rows(results, result_columns, saramin_clean_options)
                    display_rows = clean_rows
                    display_columns = clean_columns
                    if clean_saramin_enabled:
                        # 체크한 경우: 원본 결과 시트 + 정제결과 시트를 함께 저장합니다.
                        clean_sheet_added = True
                    else:
                        # 체크하지 않은 경우: 선택한 정제를 결과 시트에 바로 적용하여 한 시트로 저장합니다.
                        output_rows = clean_rows
                        output_columns = clean_columns
                        clean_rows = None
                        clean_columns = None

                excel_base64 = dataframe_to_excel_base64(
                    output_rows,
                    columns=output_columns,
                    clean_rows=clean_rows,
                    clean_columns=clean_columns,
                )

                event.update({
                    "ok": True,
                    "total_input": len(search_values),
                    "total_result": len(results),
                    "columns": display_columns,
                    "results": display_rows,
                    "collect_finance": collect_finance,
                    "clean_applied": clean_applied,
                    "clean_sheet_added": clean_sheet_added,
                    "excel_base64": excel_base64,
                    "filename": "사람인_기업정보_크롤링_결과.xlsx",
                })

            yield ndjson(event)

    return Response(
        generate(),
        mimetype="application/x-ndjson; charset=utf-8",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )

if __name__ == "__main__":
    host = os.environ.get("WORKLAB_HOST", "0.0.0.0")
    port = int(os.environ.get("WORKLAB_PORT", "8000"))
    app.run(host=host, port=port, debug=False, threaded=True)
